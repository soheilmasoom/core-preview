import uuid
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Tuple

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied, BadRequest
from django.db.models import Q, Count, F, Value, CharField, Sum, IntegerField
from django.db.models.functions import Cast, TruncDate, Greatest
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from rest_framework import serializers
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListAPIView
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.response import Response

from accounting.models import TradeRevenue
from accounts.authentication import CustomTokenAuthentication
from accounts.models import TrafficSource, User
from analytics.models import ReportPermission
from analytics.utils.list import join_lists_with_first_element
from ledger.models import Trx, Wallet


def _get_report_permission(user: User, group_id: str):
    try:
        group_id = uuid.UUID(group_id)
    except ValueError:
        raise PermissionDenied

    permission = ReportPermission.objects.filter(group_id=group_id, enable=True).first()

    if not permission:
        raise PermissionDenied

    if not user.is_superuser and permission.user != user:
        raise PermissionDenied

    return permission


def parse_date(date_str: str) -> datetime:
    try:
        return datetime.strptime(date_str or '', '%Y-%m-%d').astimezone()
    except ValueError:
        raise BadRequest('Invalid Data')


def parse_int(value: str) -> int:
    try:
        return int(value)
    except ValueError:
        pass


@login_required
def request_source_analytics(request, group_id: str):
    perm = _get_report_permission(request.user, group_id)

    context = {
        'utm_source': perm.utm_source,
        'utm_medium': perm.utm_medium,
        'redirect_url': settings.HOST_URL + f'/analytics/marketing/reports/{group_id}/download/'
    }
    return render(request, 'datetime_form.html', context)


@login_required
def get_source_analytics(request, group_id: str):
    permission = _get_report_permission(request.user, group_id)

    start = parse_date(request.GET.get('start', ''))
    end = parse_date(request.GET.get('end', '')) + timedelta(days=1)
    level = max(min(parse_int(request.GET.get('level', '')) or 2, 5), 1)

    if end - start > timedelta(days=30):
        raise BadRequest('Report can export at most 30 days')

    headers, data = get_data(permission, start, end, level)

    workbook = queryset_to_workbook(headers, data)

    # create a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reports.xlsx'

    # write workbook to response
    workbook.save(response)

    return response


def get_data(permission: ReportPermission, start: datetime, end: datetime, level: int = 2) -> Tuple[list, list]:
    utms = ['utm_source', 'utm_medium', 'utm_campaign', 'utm_content', 'utm_term'][:level]
    headers = ['date', *utms, 'users', 'verified',
               'depositors']

    queryset = TrafficSource.objects.filter(
        utm_source=permission.utm_source,
        utm_medium=permission.utm_medium,
        created__range=[start, end],
        user__account__referral__isnull=True,
    ).annotate(
        date_str=Cast(TruncDate('created'), output_field=CharField())
    ).values(
        'date_str', *utms
    ).annotate(
        user_count=Count('user_id', distinct=True),
        verified_count=Count(
            'user_id',
            distinct=True,
            filter=Q(user__level_2_verify_datetime__lte=F('user__date_joined') + Value(timedelta(days=1)))
        ),
        depositor_count=Count(
            'user_id', distinct=True,
            filter=Q(user__first_fiat_deposit_date__lte=F('user__date_joined') + Value(timedelta(days=1))) |
                   Q(user__first_crypto_deposit_date__lte=F('user__date_joined') + Value(timedelta(days=1)))
        )
    ).values_list(
        'date_str', *utms, 'user_count', 'verified_count', 'depositor_count',
    )

    data = list(queryset)

    if permission.referral_percent_revenue:
        group_by = ['date_str'] + [f'account__user__traffic_source__{s}' for s in utms]

        revenues = TradeRevenue.objects.filter(
            created__range=[start, end],
            account__user__traffic_source__utm_source=permission.utm_source,
            account__user__traffic_source__utm_medium=permission.utm_medium,
            # account__user__account__referral__isnull=True
        ).annotate(
            date_str=Cast(TruncDate('created'), output_field=CharField())
        ).values(*group_by).annotate(
            revenue=Greatest(
                Cast(
                    Sum((F('fee_revenue') + F('gap_revenue')) * F('value_irt') / F('value')) * permission.referral_percent_revenue / 100,
                    output_field=IntegerField()
                ), 0)
        ).values_list(*group_by, 'revenue')

        data = join_lists_with_first_element(data, list(revenues), n1=len(headers), n2=len(group_by) + 1,
                                             group_len=len(group_by))
        headers.append('revenue')

    return headers, sorted(list(data))


def queryset_to_workbook(headers: list, data: list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'

    # write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    # write data
    for row_num, row in enumerate(data, 1):
        for col_num, field_name in enumerate(headers, 1):
            cell = sheet.cell(row=row_num+1, column=col_num)
            cell.value = row[col_num-1]

    return workbook


class WalletSerializer(serializers.ModelSerializer):
    symbol = serializers.CharField(source='asset.symbol')
    user_id = serializers.CharField(source='account.user_id')

    class Meta:
        model = Wallet
        fields = ['id', 'created', 'user_id', 'account_id', 'market', 'symbol']


class TransactionSerializer(serializers.ModelSerializer):
    class Meta:
        model = Trx
        fields = ['id', 'created', 'sender_id', 'receiver_id', 'amount', 'group_id']


class NoCountLimitOffsetPagination(LimitOffsetPagination):
    def paginate_queryset(self, queryset, request, view=None):
        self.limit = self.get_limit(request)
        if self.limit is None:
            return None

        self.offset = self.get_offset(request)
        self.request = request

        paginated_queryset = queryset[self.offset:self.offset + self.limit]

        if not paginated_queryset:
            return []

        if self.template is not None and len(paginated_queryset) == self.limit:
            self.display_page_controls = True

        return paginated_queryset


class TransactionView(ListAPIView):
    authentication_classes = [CustomTokenAuthentication]
    pagination_class = NoCountLimitOffsetPagination
    serializer_class = TransactionSerializer

    def get_queryset(self):
        return Trx.objects.filter()

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)

        serializer = self.get_serializer(page, many=True)

        return Response({
            'results': serializer.data
        })


class WalletView(ListAPIView):
    authentication_classes = [CustomTokenAuthentication]
    pagination_class = NoCountLimitOffsetPagination
    serializer_class = WalletSerializer

    def get_queryset(self):
        return Wallet.objects.filter().prefetch_related('asset__symbol', 'account__user_id')

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)

        serializer = self.get_serializer(page, many=True)

        return Response({
            'results': serializer.data
        })